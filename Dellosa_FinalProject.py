import tkinter as tk
import openpyxl as op
from tkinter import messagebox, ttk
from datetime import datetime
import os

# ==========================================
# BACKEND FUNCTIONS
# ==========================================

def create_excel_db():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists("Dellosa_Database.xlsx"):
        wb = op.Workbook()
        ws = wb.active
        ws.title = "CarwashRecords"
        # Headers: ID, Customer, Phone, Vehicle, Plate, Service, Price, Status, Date
        ws.append(["ID", "Customer Name", "Phone", "Vehicle Type", 
                   "License Plate", "Service Type", "Price", "Status", "Date"])
        
        # Bold header styling
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        
        wb.save("Dellosa_Database.xlsx")

def display_excel():
    workbook = op.load_workbook("Dellosa_Database.xlsx")
    sheet = workbook.active

    # Clear Treeview
    for content in table.get_children():
        table.delete(content)

    # Insert Excel Data (skip header row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # Only add rows that have an ID
            table.insert("", tk.END, values=row)

def select_record(event=None):
    selected = table.focus()
    if not selected:
        return
    
    values = table.item(selected, "values")
    
    if not values:
        return
    
    # Clear current entries
    id_entry.config(state='normal')
    id_entry.delete(0, tk.END)
    customer_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    vehicle_combo.set("")
    plate_entry.delete(0, tk.END)
    service_combo.set("")
    price_entry.delete(0, tk.END)
    status_combo.set("")
    date_entry.delete(0, tk.END)
    
    # Auto-populate from selected row
    # Order: [ID, Customer, Phone, Vehicle, Plate, Service, Price, Status, Date]
    id_entry.insert(0, values[0])
    id_entry.config(state='readonly')       # ID should not be editable
    customer_entry.insert(0, values[1])
    phone_entry.insert(0, values[2])
    vehicle_combo.set(values[3])
    plate_entry.insert(0, values[4])
    service_combo.set(values[5])
    price_entry.insert(0, values[6])
    status_combo.set(values[7])
    date_entry.insert(0, values[8])

def update_price(event=None):
    """Auto-fill price when service type is selected"""
    service_prices = {
        "Basic Wash": "15.00",
        "Premium Wash": "30.00",
        "Interior Cleaning": "25.00",
        "Full Detailing": "80.00",
        "Engine Cleaning": "45.00",
        "Polish & Wax": "50.00"
    }
    selected_service = service_combo.get()
    if selected_service in service_prices:
        price_entry.config(state='normal')
        price_entry.delete(0, tk.END)
        price_entry.insert(0, service_prices[selected_service])
        price_entry.config(state='readonly')

def validation():
    customer = customer_entry.get().strip()
    phone = phone_entry.get().strip()
    vehicle = vehicle_combo.get()
    plate = plate_entry.get().strip()
    service = service_combo.get()
    status = status_combo.get()

    if not customer or not phone or not vehicle or not plate or not service or not status:
        messagebox.showerror("Error", "All fields are required!")
        return False
    
    if not phone.replace("-", "").replace(" ", "").isdigit():
        messagebox.showerror("Error", "Phone must contain only numbers!")
        return False
    
    return True

def generate_id():
    """Generate automatic ID like CW001, CW002..."""
    workbook = op.load_workbook("Dellosa_Database.xlsx")
    sheet = workbook.active
    last_row = sheet.max_row
    
    if last_row == 1:  # Only headers exist
        return "CW001"
    
    last_id = sheet.cell(row=last_row, column=1).value
    if last_id:
        num = int(last_id[2:]) + 1
        return f"CW{num:03d}"
    return "CW001"

def append_excel():
    if not validation():
        return
    
    customer = customer_entry.get().strip().title()
    phone = phone_entry.get().strip()
    vehicle = vehicle_combo.get()
    plate = plate_entry.get().strip().upper()
    service = service_combo.get()
    price = price_entry.get()
    status = status_combo.get()
    date = date_entry.get().strip()
    
    # Use current date/time if empty
    if not date:
        date = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    new_id = generate_id()
    
    workbook = op.load_workbook("Dellosa_Database.xlsx")
    sheet = workbook.active
    
    # Check duplicate license plate
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[4] == plate:  # Column index 4 = License Plate
            messagebox.showerror("Error", "This license plate already exists!")
            return
    
    sheet.append([new_id, customer, phone, vehicle, plate, service, price, status, date])
    workbook.save("Dellosa_Database.xlsx")
    
    messagebox.showinfo("Success", f"Record added successfully!\nID: {new_id}")
    clear_fields()
    display_excel()

def update_data():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Please select a record to update.")
        return
    
    values = table.item(selected, "values")
    record_id = values[0]
    
    if not validation():
        return
    
    confirm = messagebox.askyesno("Confirm Update", "Are you sure you want to update this record?")
    if not confirm:
        return
    
    customer = customer_entry.get().strip().title()
    phone = phone_entry.get().strip()
    vehicle = vehicle_combo.get()
    plate = plate_entry.get().strip().upper()
    service = service_combo.get()
    price = price_entry.get()
    status = status_combo.get()
    date = date_entry.get().strip()
    
    if not date:
        date = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    workbook = op.load_workbook("Dellosa_Database.xlsx")
    sheet = workbook.active
    
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == record_id:
            row[1].value = customer      # Customer Name
            row[2].value = phone         # Phone
            row[3].value = vehicle       # Vehicle Type
            row[4].value = plate         # License Plate
            row[5].value = service       # Service Type
            row[6].value = price         # Price
            row[7].value = status        # Status
            row[8].value = date          # Date
            found = True
            break
    
    if found:
        workbook.save("Dellosa_Database.xlsx")
        messagebox.showinfo("Success", "Record updated successfully!")
        clear_fields()
        display_excel()
    else:
        messagebox.showerror("Error", "Record not found.")

def delete_data():
    selected = table.focus()
    if not selected:
        messagebox.showerror("Error", "Please select a record to delete.")
        return
    
    values = table.item(selected, "values")
    record_id = values[0]
    
    confirm = messagebox.askyesno("Confirm Delete", 
                                  f"Are you sure you want to delete record ID {record_id}?\nThis cannot be undone!")
    if not confirm:
        return
    
    workbook = op.load_workbook("Dellosa1_Database.xlsx")
    sheet = workbook.active
    
    deleted = False
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row[0].value == record_id:
            sheet.delete_rows(i, 1)
            deleted = True
            break
    
    if deleted:
        workbook.save("Dellosa_Database.xlsx")
        messagebox.showinfo("Success", "Record deleted successfully!")
        clear_fields()
        display_excel()
    else:
        messagebox.showerror("Error", "Record not found.")

def clear_fields():
    id_entry.config(state='normal')
    id_entry.delete(0, tk.END)
    id_entry.config(state='readonly')
    customer_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    vehicle_combo.set("")
    plate_entry.delete(0, tk.END)
    service_combo.set("")
    price_entry.config(state='normal')
    price_entry.delete(0, tk.END)
    price_entry.config(state='readonly')
    status_combo.set("")
    date_entry.delete(0, tk.END)




window = tk.Tk()
window.title("Car Wash Service Management System")
window.configure(bg="black")
window.geometry("950x650")

# Create DB on first run
create_excel_db()

# Title
title = tk.Label(window, text="Car Wash Service Management System", 
                 font=("Times New Roman", 16, "bold"), bg="gold")
title.grid(row=0, column=0, columnspan=8, pady=(10, 5))

# Input Frame
genframe = tk.Frame(window, bg="black", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=8, padx=10, pady=10, sticky="ew")


id_label = tk.Label(genframe, text="ID:", font=("Poppins", 10, "italic"), bg="gold")
id_label.grid(row=0, column=0, padx=(10, 0), pady=(10, 0), sticky="w")

id_entry = tk.Entry(genframe, font=("Poppins", 12), width=12, state='readonly')
id_entry.grid(row=0, column=1, padx=(5, 10), pady=(10, 0))

customer_label = tk.Label(genframe, text="Customer Name:", font=("Poppins", 10, "italic"), bg="red")
customer_label.grid(row=0, column=2, padx=(10, 0), pady=(10, 0), sticky="w")

customer_entry = tk.Entry(genframe, font=("Poppins", 12), width=22)
customer_entry.grid(row=0, column=3, columnspan=2, padx=(5, 10), pady=(10, 0))

phone_label = tk.Label(genframe, text="Phone:", font=("Poppins", 10, "italic"), bg="gold")
phone_label.grid(row=0, column=5, padx=(10, 0), pady=(10, 0), sticky="w")

phone_entry = tk.Entry(genframe, font=("Poppins", 12), width=15)
phone_entry.grid(row=0, column=6, padx=(5, 10), pady=(10, 0))


vehicle_label = tk.Label(genframe, text="Vehicle Type:", font=("Poppins", 10, "italic"), bg="red")
vehicle_label.grid(row=1, column=0, padx=(10, 0), pady=(10, 0), sticky="w")

vehicle_combo = ttk.Combobox(genframe, font=("Poppins", 12), width=19, state="readonly",
                             values=["Sedan", "SUV", "Truck", "Van", "Motorcycle"])
vehicle_combo.grid(row=1, column=1, padx=(5, 10), pady=(10, 0))

plate_label = tk.Label(genframe, text="License Plate:", font=("Poppins", 10, "italic"), bg="gold")
plate_label.grid(row=1, column=2, padx=(10, 0), pady=(10, 0), sticky="w")

plate_entry = tk.Entry(genframe, font=("Poppins", 12), width=22)
plate_entry.grid(row=1, column=3, columnspan=2, padx=(5, 10), pady=(10, 0))

service_label = tk.Label(genframe, text="Service Type:", font=("Poppins", 10, "italic"), bg="red")
service_label.grid(row=1, column=5, padx=(10, 0), pady=(10, 0), sticky="w")

service_combo = ttk.Combobox(genframe, font=("Poppins", 12), width=12, state="readonly",
                             values=["Basic Wash", "Premium Wash", "Interior Cleaning", 
                                     "Full Detailing", "Engine Cleaning", "Polish & Wax"])
service_combo.grid(row=1, column=6, padx=(5, 10), pady=(10, 0))
service_combo.bind("<<ComboboxSelected>>", update_price)

price_label = tk.Label(genframe, text="Price ($):", font=("Poppins", 10, "italic"), bg="gold")
price_label.grid(row=2, column=0, padx=(10, 0), pady=(10, 0), sticky="w")

price_entry = tk.Entry(genframe, font=("Poppins", 12), width=12, state='readonly')
price_entry.grid(row=2, column=1, padx=(5, 10), pady=(10, 0))

status_label = tk.Label(genframe, text="Status:", font=("Poppins", 10, "italic"), bg="red")
status_label.grid(row=2, column=2, padx=(10, 0), pady=(10, 0), sticky="w")

status_combo = ttk.Combobox(genframe, font=("Poppins", 12), width=19, state="readonly",
                            values=["Pending", "In Progress", "Completed", "Cancelled"])
status_combo.grid(row=2, column=3, columnspan=2, padx=(5, 10), pady=(10, 0))

date_label = tk.Label(genframe, text="Date:", font=("Poppins", 10, "italic"), bg="gold")
date_label.grid(row=2, column=5, padx=(10, 0), pady=(10, 0), sticky="w")

date_entry = tk.Entry(genframe, font=("Poppins", 12), width=15)
date_entry.grid(row=2, column=6, padx=(5, 10), pady=(10, 0))


date_entry.insert(0, datetime.now().strftime("%Y-%m-%d %H:%M"))


btn_frame = tk.Frame(window, bg="red")
btn_frame.grid(row=2, column=0, columnspan=8, pady=(5, 10))

submit_button = tk.Button(btn_frame, text="➕ Add Record", font=("Poppins", 12, "bold"),
                          bg="red", command=append_excel, width=14)
submit_button.grid(row=0, column=0, padx=5)

update_btn = tk.Button(btn_frame, text="✏️ Update", font=("Poppins", 12, "bold"),
                       bg="gold", command=update_data, width=14)
update_btn.grid(row=0, column=1, padx=5)

delete_btn = tk.Button(btn_frame, text="️ Delete", font=("Poppins", 12, "bold"),
                       bg="gold", fg="black", command=delete_data, width=14)
delete_btn.grid(row=0, column=2, padx=5)

clear_btn = tk.Button(btn_frame, text="🔄 Clear", font=("Poppins", 12, "bold"),
                      bg="gold", fg="black", command=clear_fields, width=14)
clear_btn.grid(row=0, column=3, padx=5)

refresh_btn = tk.Button(btn_frame, text="📊 Refresh", font=("Poppins", 12, "bold"),
                        bg="red", command=display_excel, width=14)
refresh_btn.grid(row=0, column=4, padx=5)


columns = ("ID", "Customer Name", "Phone", "Vehicle Type", 
           "License Plate", "Service Type", "Price", "Status", "Date")

table = ttk.Treeview(window, columns=columns, show="headings")

for col in columns:
    table.heading(col, text=col)
    if col == "ID":
        table.column(col, width=70, anchor=tk.CENTER)
    elif col == "Customer Name":
        table.column(col, width=130, anchor=tk.W)
    elif col == "Phone":
        table.column(col, width=100, anchor=tk.CENTER)
    elif col == "Vehicle Type":
        table.column(col, width=90, anchor=tk.CENTER)
    elif col == "License Plate":
        table.column(col, width=100, anchor=tk.CENTER)
    elif col == "Service Type":
        table.column(col, width=110, anchor=tk.W)
    elif col == "Price":
        table.column(col, width=70, anchor=tk.CENTER)
    elif col == "Status":
        table.column(col, width=90, anchor=tk.CENTER)
    else:
        table.column(col, width=130, anchor=tk.CENTER)

table.grid(row=3, column=0, columnspan=8, padx=10, pady=(0, 10), sticky="nsew")

# Bind selection event
table.bind("<<TreeviewSelect>>", select_record)

# Configure grid weight so treeview expands
window.grid_rowconfigure(3, weight=1)
window.grid_columnconfigure(0, weight=1)

# Load data on startup
display_excel()

window.mainloop()